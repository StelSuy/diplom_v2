"""Export route — CSV/XLSX reports."""
import csv
import io
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import require_admin
from app.db.session import get_db
from app.models.user import User
from app.models.employee import Employee
from app.models.event import Event
from app.services.worktime import build_intervals, hms_from_seconds, split_interval_seconds_by_local_day
from app.core.time import WARSAW

router = APIRouter(prefix="/export", tags=["export"])


def _build_report(db: Session, date_from: date, date_to: date, employee_id: Optional[int] = None):
    q = db.query(Employee).filter(Employee.is_active == True)
    if employee_id:
        q = q.filter(Employee.id == employee_id)
    employees = q.order_by(Employee.full_name).all()

    from datetime import datetime, timezone
    dt_from = datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc)
    dt_to = datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc)

    rows = []
    for emp in employees:
        events = (
            db.query(Event)
            .filter(Event.employee_id == emp.id, Event.ts >= dt_from, Event.ts <= dt_to)
            .order_by(Event.ts)
            .all()
        )
        intervals, _, _ = build_intervals(events)
        total_seconds = sum(
            int((iv.out_utc - iv.in_utc).total_seconds())
            for iv in intervals
        )
        worked_days = set()
        for iv in intervals:
            buckets = split_interval_seconds_by_local_day(iv.in_utc, iv.out_utc)
            for day_str in buckets:
                if buckets[day_str] > 0:
                    worked_days.add(day_str)

        rows.append({
            "id": emp.id,
            "full_name": emp.full_name,
            "position": emp.position or "",
            "total_hms": hms_from_seconds(total_seconds),
            "total_minutes": total_seconds // 60,
            "worked_days": len(worked_days),
        })
    return rows


@router.get("/worktime.csv")
def export_csv(
    date_from: date = Query(...),
    date_to: date = Query(...),
    employee_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    if date_from > date_to:
        raise HTTPException(400, "date_from > date_to")

    rows = _build_report(db, date_from, date_to, employee_id)

    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM for Excel
    writer = csv.DictWriter(output, fieldnames=["id", "full_name", "position", "total_hms", "total_minutes", "worked_days"], delimiter=";")
    writer.writeheader()
    for r in rows:
        writer.writerow(r)

    output.seek(0)
    filename = f"worktime_{date_from}_{date_to}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/worktime.xlsx")
def export_xlsx(
    date_from: date = Query(...),
    date_to: date = Query(...),
    employee_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    if date_from > date_to:
        raise HTTPException(400, "date_from > date_to")

    rows = _build_report(db, date_from, date_to, employee_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worktime"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["ID", "ПІБ", "Посада", "Відпрацьовано (год:хв:с)", "Хвилин", "Робочих днів"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r["id"])
        ws.cell(row=row_idx, column=2, value=r["full_name"])
        ws.cell(row=row_idx, column=3, value=r["position"])
        ws.cell(row=row_idx, column=4, value=r["total_hms"])
        ws.cell(row=row_idx, column=5, value=r["total_minutes"])
        ws.cell(row=row_idx, column=6, value=r["worked_days"])

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"worktime_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
